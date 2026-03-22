import csv
from datetime import date, datetime
from io import BytesIO, StringIO
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from .errors import APIError
from .extensions import db
from .models import AttachmentInfo, ExportTask
from .platform_ops import build_storage_subdir, export_dir_path
from .time_utils import now_beijing


XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def format_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return value


def create_workbook_bytes(sheets):
    workbook = Workbook()
    header_fill = PatternFill("solid", fgColor="1f6feb")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for index, sheet in enumerate(sheets):
        worksheet = workbook.active if index == 0 else workbook.create_sheet()
        worksheet.title = (sheet.get("title") or f"Sheet{index + 1}")[:31]
        headers = sheet.get("headers") or []
        rows = sheet.get("rows") or []

        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for row in rows:
            worksheet.append([format_cell_value(item) for item in row])

        for column_cells in worksheet.columns:
            max_length = 0
            for cell in column_cells:
                cell_value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(cell_value))
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 40)

        worksheet.freeze_panes = "A2"

    content = BytesIO()
    workbook.save(content)
    return content.getvalue()


def normalize_header(value):
    return str(value).strip() if value not in (None, "") else ""


def normalize_attachment_name(raw_name, default_ext=""):
    filename = str(raw_name or "").replace("\\", "/").split("/")[-1].strip()
    if filename:
        return filename
    return f"{uuid4().hex}{default_ext}"


def build_safe_storage_name(display_name, default_ext=""):
    normalized_name = normalize_attachment_name(display_name, default_ext)
    original_path = Path(normalized_name)
    suffix = original_path.suffix or default_ext
    safe_stem = secure_filename(original_path.stem)
    if not safe_stem:
        safe_stem = uuid4().hex
    safe_suffix = secure_filename(str(suffix or "")).lower() or str(suffix or "").lower()
    if safe_suffix and not safe_suffix.startswith("."):
        safe_suffix = f".{safe_suffix}"
    return normalized_name, f"{uuid4().hex}_{safe_stem}{safe_suffix}"


def parse_tabular_file(file_storage):
    if not isinstance(file_storage, FileStorage):
        raise APIError("上传文件格式不合法")
    filename = (file_storage.filename or "").lower()
    if not filename:
        raise APIError("上传文件不能为空")

    if filename.endswith(".csv"):
        text = file_storage.stream.read().decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text))
        return [{normalize_header(key): value for key, value in row.items()} for row in reader]

    workbook = load_workbook(file_storage.stream, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize_header(item) for item in rows[0]]
    records = []
    for row in rows[1:]:
        if not any(item not in (None, "") for item in row):
            continue
        payload = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            payload[header] = row[index] if index < len(row) else None
        records.append(payload)
    return records


def save_uploaded_attachment(file_storage, biz_type, uploader_id, subdir="imports"):
    filename, stored_name = build_safe_storage_name(file_storage.filename)

    target_dir = build_storage_subdir(subdir)
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / stored_name
    file_storage.save(target_path)

    attachment = AttachmentInfo(
        file_name=filename,
        file_path=str(target_path),
        file_ext=(Path(filename).suffix or target_path.suffix).lower().lstrip("."),
        file_size=target_path.stat().st_size,
        biz_type=biz_type,
        uploader_id=uploader_id,
    )
    db.session.add(attachment)
    db.session.flush()
    return attachment


def create_binary_attachment(filename, content, biz_type, creator_id, subdir="exports"):
    target_dir = export_dir_path() if subdir == "exports" else build_storage_subdir(subdir)
    target_dir.mkdir(parents=True, exist_ok=True)
    display_name, stored_name = build_safe_storage_name(filename, Path(str(filename or "")).suffix or "")
    target_path = target_dir / stored_name
    target_path.write_bytes(content)

    attachment = AttachmentInfo(
        file_name=display_name,
        file_path=str(target_path),
        file_ext=(Path(display_name).suffix or target_path.suffix).lower().lstrip("."),
        file_size=target_path.stat().st_size,
        biz_type=biz_type,
        uploader_id=creator_id,
    )
    db.session.add(attachment)
    db.session.flush()
    return attachment, target_path


def create_export_file(filename, content, biz_type, creator_id, create_task=True):
    attachment, target_path = create_binary_attachment(filename, content, biz_type, creator_id, subdir="exports")
    display_name = attachment.file_name

    task = None
    if create_task:
        finished_at = now_beijing()
        task = ExportTask(
            task_type=biz_type,
            task_name=Path(display_name).stem,
            file_id=attachment.id,
            status="completed",
            progress=100,
            current_step="导出完成",
            error_message="",
            request_payload_json="{}",
            started_at=finished_at,
            finished_at=finished_at,
            retry_count=0,
            created_by=creator_id,
        )
        db.session.add(task)
        db.session.flush()
    return attachment, task, target_path
